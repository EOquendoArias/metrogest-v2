#!/usr/bin/env python3
"""
generar_excel_prueba_fase5.py — genera un archivo .xlsx con las 6 hojas
nuevas de la Fase 5 (docs/migracion/PLAN_FASE5_EXTENSIONES.md): Planes-
Verificacion, Verificaciones, PuntosVerificacion, Evaluaciones, Planes-
Mantenimiento, Mantenimientos — con errores DELIBERADOS, uno por cada
regla nueva de esa fase (además de las 4 hojas base del MVP, limpias, para
no meter ruido en el conteo).

Uso:
    python generar_excel_prueba_fase5.py [ruta_salida.xlsx]

Ver comentarios junto a cada fila para el detalle de qué error/regla/
severidad debería disparar cada una.
"""
import sys
import openpyxl

RUTA = sys.argv[1] if len(sys.argv) > 1 else "excel_prueba_fase5.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# -- Hojas base del MVP, deliberadamente limpias (sin desviaciones) --------
ws = wb.create_sheet("Equipos")
ws.append(["codigo", "nombre", "estado"])
ws.append(["EQ-F5-001", "Balanza Fase 5", "operativo"])
ws.append(["EQ-F5-002", "Termómetro Fase 5", "operativo"])

ws = wb.create_sheet("Magnitudes")
ws.append(["codigo_equipo", "nombre_magnitud", "emp_valor"])
ws.append(["EQ-F5-001", "Masa", 1.0])
ws.append(["EQ-F5-002", "Temperatura", 0.5])
ws.append(["EQ-F5-002", "Humedad", 2.0])  # existe, pero sin plan de verificación a propósito

ws = wb.create_sheet("Calibraciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha_calibracion"])
# Intervalo real entre estas dos: ~6 meses -> usado para derivar el
# intervalo_adoptado_meses de la Evaluación de EQ-F5-001/Masa (§1.3 del plan)
ws.append(["CAL-F5-001", "EQ-F5-001", "Masa", "2022-01-01"])
ws.append(["CAL-F5-002", "EQ-F5-001", "Masa", "2022-07-01"])

ws = wb.create_sheet("PuntosCalibracion")
ws.append(["id_temporal_calibracion", "numero_punto", "valor_patron", "valor_indicado"])
ws.append(["CAL-F5-001", 1, 10, 10.2])
ws.append(["CAL-F5-002", 1, 10, 10.1])

# -- Hojas de Fase 5, con errores deliberados -------------------------------

ws = wb.create_sheet("PlanesVerificacion")
ws.append(["codigo_equipo", "nombre_magnitud", "frecuencia_meses", "justificacion_no_aplica"])
# Error 1: sin frecuencia ni justificación -> Media (nivel2-plan-sin-frecuencia-ni-justificacion)
ws.append(["EQ-F5-001", "Masa", None, None])
ws.append(["EQ-F5-002", "Temperatura", 6, None])
# Error 2: magnitud inexistente (EQ-F5-999 no existe) -> Crítica (nivel3-referencial-magnitud-inexistente)
ws.append(["EQ-F5-999", "Presion", None, None])
# Error 3: duplicado de la fila 2 (EQ-F5-002/Temperatura) -> Crítica (capa1-duplicado-dentro-del-archivo)
ws.append(["EQ-F5-002", "Temperatura", 6, None])

ws = wb.create_sheet("Verificaciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha"])
ws.append(["VER-F5-001", "EQ-F5-002", "Temperatura", "2023-01-01"])  # plan sí existe
ws.append(["VER-F5-002", "EQ-F5-001", "Masa", "2023-01-01"])  # plan existe (aunque con Media pendiente)
# Error 4: Humedad tiene magnitud válida pero SIN plan declarado -> Crítica (nivel3-referencial-plan-verificacion-inexistente)
ws.append(["VER-F5-003", "EQ-F5-002", "Humedad", "2023-01-01"])
# Error 5: id_temporal repetido (VER-F5-001 ya existe arriba) -> Crítica (capa1-duplicado-dentro-del-archivo)
ws.append(["VER-F5-001", "EQ-F5-002", "Temperatura", "2023-02-01"])

ws = wb.create_sheet("PuntosVerificacion")
ws.append(["id_temporal_verificacion", "numero_punto", "valor_patron", "valor_indicado"])
ws.append(["VER-F5-001", 1, 10, 10.1])
# Error 6: referencia a una verificación que no existe -> Crítica (nivel3-referencial-verificacion-inexistente)
ws.append(["VER-F5-999", 1, 10, 10.1])

ws = wb.create_sheet("Evaluaciones")
ws.append(["codigo_equipo", "nombre_magnitud", "intervalo_adoptado_meses", "justificacion_exceso"])
# Fila limpia: sin ado declarado -> se deriva del historial real (~6 meses,
# ver Calibraciones arriba); sugerido con factores por defecto (todos 3) = 12
# -> 6 <= 12, sin exceso, sin desviación
ws.append(["EQ-F5-001", "Masa", None, None])
# Error 7: ado=24 declarado > sugerido (12 con factores por defecto), sin
# justificacion_exceso -> Alta (nivel4-ilac-adoptado-excede-sugerido-sin-justificacion)
ws.append(["EQ-F5-002", "Temperatura", 24, None])
# Error 8: magnitud inexistente -> Crítica (nivel3-referencial-magnitud-inexistente)
ws.append(["EQ-F5-999", "Presion", None, None])
# Error 9: duplicado de la fila 1 (EQ-F5-001/Masa) -> Crítica (capa1-duplicado-dentro-del-archivo)
ws.append(["EQ-F5-001", "Masa", None, None])

ws = wb.create_sheet("PlanesMantenimiento")
ws.append(["codigo_equipo", "tipo", "frecuencia_meses"])
ws.append(["EQ-F5-001", "preventivo", 6])
# Error 10: equipo inexistente -> Crítica (nivel3-referencial-equipo-inexistente)
ws.append(["EQ-F5-999", "preventivo", 6])

ws = wb.create_sheet("Mantenimientos")
ws.append(["codigo_equipo", "tipo", "origen", "titulo"])
ws.append(["EQ-F5-001", "correctivo", "interno", "Cambio de sensor"])
# Error 11: 'tipo' obligatorio vacío -> Crítica (nivel1-estructural-valor-obligatorio-vacio)
ws.append(["EQ-F5-002", None, "interno", "Ajuste"])
# Error 12: equipo inexistente -> Crítica (nivel3-referencial-equipo-inexistente)
ws.append(["EQ-F5-999", "preventivo", "externo", "Revisión anual"])

wb.save(RUTA)
print(f"Archivo de prueba Fase 5 generado: {RUTA}")
print("Esperado (--dry-run, sin BD): 10 Crítica, 1 Alta, 2 Media, 0 Baja")
print("(la fila EQ-F5-999 de PlanesVerificacion dispara 2 reglas independientes "
      "a la vez: magnitud inexistente Y sin frecuencia/justificación)")
print("(las hojas base MVP deben salir limpias - 0 desviaciones ahí)")
