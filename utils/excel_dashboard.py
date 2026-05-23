"""Excel del Dashboard — tabla detallada de equipos."""
import io
from datetime import date, datetime

def generar_excel_dashboard(datos, hoy, config):
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Equipos"

    nombre_lab = getattr(config,'nombre','MetroGest') or 'MetroGest'

    # Estilos
    AZ   = "0F3460"; AZ2 = "1A4F8A"; OK = "16A34A"; FL = "DC2626"; AM = "D97706"
    BG   = "F8FAFC"; GD  = "E2E8F0"

    def st_header():
        return Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def border():
        s = Side(style='thin', color=GD)
        return Border(left=s, right=s, top=s, bottom=s)
    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f"{nombre_lab} — Estado de Equipos — {hoy.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color=AZ2)
    ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', italic=True, size=9, color='64748B')
    ws['A2'].alignment = center()

    # Indicadores en fila 3
    ws.merge_cells('A3:B3'); ws['A3'] = "Total equipos"
    ws.merge_cells('C3:D3'); ws['C3'] = "Operativos"
    ws.merge_cells('E3:F3'); ws['E3'] = "Calibraciones vencidas"
    ws.merge_cells('G3:H3'); ws['G3'] = f"Costo calibraciones {hoy.year}"
    ws['A3'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['C3'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['E3'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['G3'].font = Font(bold=True, color='FFFFFF', size=9)
    for c in ['A3','B3','C3','D3','E3','F3','G3','H3']:
        ws[c].fill = fill(AZ2)
        ws[c].alignment = center()

    ws.merge_cells('A4:B4'); ws['A4'] = datos['total']
    ws.merge_cells('C4:D4'); ws['C4'] = datos['operativos']
    ws.merge_cells('E4:F4'); ws['E4'] = datos['vencidos']
    ws.merge_cells('G4:H4'); ws['G4'] = datos['costo_anio']
    ws['A4'].font = Font(bold=True, size=14, color=AZ2)
    ws['C4'].font = Font(bold=True, size=14, color=OK)
    ws['E4'].font = Font(bold=True, size=14, color=FL)
    ws['G4'].font = Font(bold=True, size=12, color=AZ)
    ws['G4'].number_format = '"$"#,##0.00'
    for c in ['A4','C4','E4','G4']:
        ws[c].alignment = center()

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 24

    # Encabezado tabla
    headers = ['Equipo','Código','Área','Estado','Responsable',
               'Magnitudes activas','Última calibración','Próxima calibración',
               'Días para vencer','Último mantenimiento']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = st_header()
        cell.fill = fill(AZ2)
        cell.alignment = center()
        cell.border = border()
    ws.row_dimensions[6].height = 22

    # Datos
    for row_idx, d in enumerate(datos['filas'], 7):
        eq = d['equipo']
        dias = d['dias_cal']

        if dias is None:
            dias_txt = 'Sin programar'
        elif dias < 0:
            dias_txt = f'Vencida {abs(dias)}d'
        else:
            dias_txt = f'{dias}d'

        fila = [
            eq.nombre,
            eq.codigo,
            eq.area or '—',
            eq.estado.replace('_',' ').title(),
            eq.responsable or '—',
            len([m for m in eq.magnitudes if m.activa]),
            d['ultima_cal'].strftime('%d/%m/%Y') if d['ultima_cal'] else '—',
            d['proxima_cal'].strftime('%d/%m/%Y') if d['proxima_cal'] else '—',
            dias_txt,
            d['ultimo_mant'].strftime('%d/%m/%Y') if d['ultimo_mant'] else '—',
        ]
        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border()
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.font = Font(name='Calibri', size=9)
            if row_idx % 2 == 0:
                cell.fill = fill(BG)

        # Color días
        dias_cell = ws.cell(row=row_idx, column=9)
        if dias is not None:
            if dias < 0:
                dias_cell.font = Font(name='Calibri', size=9, bold=True, color=FL)
            elif dias <= 30:
                dias_cell.font = Font(name='Calibri', size=9, bold=True, color=AM)

    # Anchos de columna
    anchos = [30, 10, 15, 18, 18, 10, 16, 16, 14, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Hoja de áreas
    ws2 = wb.create_sheet("Equipos por Área")
    ws2['A1'] = "Área"; ws2['B1'] = "Cantidad"
    ws2['A1'].font = Font(bold=True, color='FFFFFF')
    ws2['B1'].font = Font(bold=True, color='FFFFFF')
    ws2['A1'].fill = fill(AZ2); ws2['B1'].fill = fill(AZ2)
    ws2['A1'].alignment = center(); ws2['B1'].alignment = center()
    for r, (area, cant) in enumerate(datos['por_area'].items(), 2):
        ws2[f'A{r}'] = area; ws2[f'B{r}'] = cant
        ws2[f'B{r}'].alignment = center()
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_excel_listado(datos, hoy, config):
    """Excel del listado general de equipos (auditoría)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado General"

    nombre_lab = getattr(config, 'nombre', 'MetroGest') or 'MetroGest'
    AZ2 = "1A4F8A"; FL = "DC2626"; AM = "D97706"; OK = "16A34A"
    BG  = "F8FAFC"; GD = "E2E8F0"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def border():
        s = Side(style='thin', color=GD)
        return Border(left=s, right=s, top=s, bottom=s)
    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def hdr_font():
        return Font(name='Calibri', bold=True, color='FFFFFF', size=9)

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f"{nombre_lab} — Listado General de Equipos — {hoy.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color=AZ2)
    ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generado: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', italic=True, size=9, color='64748B')
    ws['A2'].alignment = center()

    # Encabezados
    headers = [
        'Equipo', 'Código', 'Estado', 'Última cal.',
        'Próxima cal.', 'Días cal.', 'Verif. estado', 'Próxima verif.', 'Último mant.', 'Área'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font()
        cell.fill = fill(AZ2)
        cell.alignment = center()
        cell.border = border()
    ws.row_dimensions[4].height = 20

    verif_labels = {
        'al_dia':   'Al día',
        'pendiente':'Pendiente',
        'vencida':  'Vencida',
        'sin_plan': 'Sin plan',
        'no_aplica':'No aplica',
    }

    for row_idx, d in enumerate(datos, 5):
        eq = d['equipo']
        dias = d['dias_cal']
        if dias is None:
            dias_txt = 'Sin programar'
        elif dias < 0:
            dias_txt = f'Vencida {abs(dias)}d'
        elif dias == 0:
            dias_txt = 'Hoy'
        else:
            dias_txt = f'{dias}d'

        ve = d.get('verif_estado', '')
        vp = d.get('verif_proxima')

        fila = [
            eq.nombre,
            eq.codigo,
            eq.estado.replace('_', ' ').title(),
            d['ultima_cal'].strftime('%d/%m/%Y') if d['ultima_cal'] else '—',
            d['proxima_cal'].strftime('%d/%m/%Y') if d['proxima_cal'] else '—',
            dias_txt,
            verif_labels.get(ve, '—'),
            vp.strftime('%d/%m/%Y') if vp else '—',
            d['ultimo_mant'].strftime('%d/%m/%Y') if d['ultimo_mant'] else '—',
            eq.area or '—',
        ]
        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border()
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = fill(BG)

        # Color días cal.
        dias_cell = ws.cell(row=row_idx, column=6)
        if dias is not None:
            if dias < 0:
                dias_cell.font = Font(name='Calibri', size=9, bold=True, color=FL)
            elif dias <= 30:
                dias_cell.font = Font(name='Calibri', size=9, bold=True, color=AM)

        # Color verif. estado
        ve_cell = ws.cell(row=row_idx, column=7)
        if ve == 'vencida':
            ve_cell.font = Font(name='Calibri', size=9, bold=True, color=FL)
        elif ve == 'pendiente':
            ve_cell.font = Font(name='Calibri', size=9, bold=True, color=AM)
        elif ve == 'al_dia':
            ve_cell.font = Font(name='Calibri', size=9, color=OK)

    # Anchos de columna
    anchos = [28, 12, 18, 14, 14, 14, 14, 14, 14, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
