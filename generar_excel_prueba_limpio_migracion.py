#!/usr/bin/env python3
"""
generar_excel_prueba_limpio_migracion.py — genera un archivo .xlsx en la
plantilla estándar (docs/migracion/PLAN_IMPORTACION_EXCEL.md §3) SIN
ningún error deliberado — 2 equipos, 2 magnitudes, 2 calibraciones, 4
puntos, todos dentro de tolerancia y con `resultado` coincidiendo con el
semáforo recalculado.

Complementa a generar_excel_prueba_migracion.py (que sí tiene errores
deliberados, para probar que el importador los detecta). Este archivo
limpio sirve para probar el CAMINO POSITIVO: que --ejecutar realmente
inserta filas correctas en la base de datos — ver el runbook en
docs/migracion/PLAN_IMPORTACION_EXCEL.md §9.

Uso:
    python generar_excel_prueba_limpio_migracion.py [ruta_salida.xlsx] [sufijo]

El [sufijo] (ej. "-R2") se agrega a los códigos EQ-CLEAN-001/002 y a los
certificados CERT-CLEAN-1001/1002, para poder repetir la prueba de carga
real sin chocar con equipos que ya quedaron insertados en una corrida
anterior (la Capa 2 de duplicados los saltaría silenciosamente y no
serviría para probar una inserción nueva).
"""
import sys
import openpyxl

RUTA = sys.argv[1] if len(sys.argv) > 1 else "excel_prueba_limpio.xlsx"
SUF = sys.argv[2] if len(sys.argv) > 2 else ""

EQ1, EQ2 = f"EQ-CLEAN-001{SUF}", f"EQ-CLEAN-002{SUF}"
CAL1, CAL2 = f"CAL-CLEAN-001{SUF}", f"CAL-CLEAN-002{SUF}"
CERT1, CERT2 = f"CERT-CLEAN-1001{SUF}", f"CERT-CLEAN-1002{SUF}"
SN1, SN2 = f"SN-CLEAN-001{SUF}", f"SN-CLEAN-002{SUF}"

wb = openpyxl.Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("Equipos")
ws.append(["codigo", "nombre", "descripcion", "marca", "modelo", "numero_serie",
           "numero_inventario", "fecha_adquisicion", "costo", "area", "ubicacion",
           "responsable", "estado"])
ws.append([EQ1, "Balanza prueba limpia", "", "Mettler Toledo", "XS205",
           SN1, "", "2020-01-15", 8500000, "Laboratorio 1", "Mesa 3",
           "Juan Pérez", "operativo"])
ws.append([EQ2, "Termómetro prueba limpia", "", "Fluke", "1523",
           SN2, "", "2019-06-01", 1200000, "Laboratorio 1", "Estante A",
           "Ana Gómez", "operativo"])

ws = wb.create_sheet("Magnitudes")
ws.append(["codigo_equipo", "nombre_magnitud", "simbolo", "unidad", "rango_min",
           "rango_max", "resolucion", "emp_texto", "emp_valor", "emp_unidad",
           "clase_exactitud", "tipo_instrumento"])
ws.append([EQ1, "Masa", "m", "g", 0, 220, "0.0001", "±1.0 g", 1.0, "g",
           "I", "continuo"])
ws.append([EQ2, "Temperatura", "T", "°C", -20, 100, "0.1", "±0.5 °C", 0.5,
           "°C", "", "continuo"])

ws = wb.create_sheet("Calibraciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha_calibracion",
           "numero_certificado", "laboratorio", "acreditacion_laboratorio",
           "proxima_calibracion", "patrones_utilizados", "metodo_calibracion",
           "temperatura_ambiente", "humedad_relativa", "trazabilidad",
           "observaciones", "costo", "resultado"])
# Puntos dentro de EMP (1.0 g) -> recalculado = "aprobado", coincide con lo declarado
ws.append([CAL1, EQ1, "Masa", "2023-01-15", CERT1,
           "Lab Metrología SAS", "ONAC", "2024-01-15", "Pesas patrón clase E2",
           "Comparación directa", 22.5, 45, "Trazable a INM", "", 250000, "aprobado"])
# Puntos dentro de EMP (0.5 °C) -> recalculado = "aprobado", coincide con lo declarado
ws.append([CAL2, EQ2, "Temperatura", "2023-02-10", CERT2,
           "Lab Térmico Ltda", "ONAC", "2024-02-10", "Baño termostático",
           "Comparación directa", 23.0, 50, "Trazable a INM", "", 180000, "aprobado"])

ws = wb.create_sheet("PuntosCalibracion")
ws.append(["id_temporal_calibracion", "numero_punto", "valor_patron", "valor_indicado",
           "incertidumbre", "observacion"])
ws.append([CAL1, 1, 10, 10.3, 0.05, ""])
ws.append([CAL1, 2, 20, 20.2, 0.05, ""])
ws.append([CAL2, 1, 25.0, 25.1, 0.05, ""])
ws.append([CAL2, 2, 50.0, 50.1, 0.05, ""])

wb.save(RUTA)
print(f"Archivo de prueba LIMPIO generado: {RUTA} (códigos con sufijo: '{SUF}')")
print("Debe dar 0 desviaciones de cualquier severidad en modo --dry-run.")
