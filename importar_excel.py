#!/usr/bin/env python3
"""
importar_excel.py — Importador de datos históricos de clientes hacia MetroGest v2

Fase 2 de docs/migracion/PLAN_IMPORTACION_EXCEL.md. Lee la PLANTILLA
ESTÁNDAR ya transformada por Power Query desde el Excel real del cliente
(ver ese documento §3 para el esquema exacto de columnas) — este script
NUNCA intenta adivinar el formato original del cliente, eso es
responsabilidad de la transformación de Power Query, no de este código.

Implementa los 5 niveles de validación y las 3 capas de detección de
duplicados de docs/migracion/PLAN_IMPORTACION_EXCEL.md §4-§5, y produce el
Registro de Desviaciones descrito en
docs/migracion/GUIA_VALIDACION_Y_DESVIACIONES.md §4.

MODELO DE SEGURIDAD — distinto del de seed_carga_masiva.py a propósito:
seed_carga_masiva.py genera datos DESECHABLES y por eso EXIGE que el
nombre de la BD contenga test/carga/staging. Este script maneja datos
REALES de un cliente que eventualmente sí van a la base de datos de
producción — por eso no bloquea por nombre de BD, pero:
  - Nunca escribe nada sin --ejecutar explícito (por defecto es --dry-run).
  - Si la URL no parece de staging, exige --confirmo-produccion a propósito.
  - Exige confirmación interactiva escrita antes de tocar la base de datos.
  - Vuelve a correr TODA la validación antes de escribir, nunca confía en
    un reporte de una corrida anterior.
  - Si quedan desviaciones Críticas, se niega a continuar sin excepción.
  - Si quedan desviaciones Altas, exige un --registro-resueltas que las
    tenga todas marcadas como Resueltas o Aceptadas (ver GUIA_VALIDACION_
    Y_DESVIACIONES.md §5) — nunca decide solo si una Alta se puede pasar.

MODO SIN BASE DE DATOS: si se corre sin --database-url, el script igual
valida estructura/tipos/referencias dentro del archivo y duplicados dentro
del propio Excel (Capa 1) — no puede validar contra la BD (Capa 2/3) ni
ejecutar la carga real. Es intencional: permite que el propio cliente (o
Edison sin acceso a la BD del cliente todavía) revise su archivo antes de
enviarlo, sin necesitar credenciales de base de datos.

Verificado (12-ago-2026) contra un Excel de prueba con 14 errores
deliberados (ver generar_excel_prueba_migracion.py): el modo --dry-run
detectó los 14 con la severidad exacta esperada. También verificado
--ejecutar contra Postgres real (metrogest_carga) — escritura, rastro de
auditoría y detección de duplicados en 3 capas, ver
docs/migracion/PLAN_IMPORTACION_EXCEL.md §9-§10.

FASE 5 (12-ago-2026, ver docs/migracion/PLAN_FASE5_EXTENSIONES.md): se
agregaron 6 hojas opcionales (PlanesVerificacion, Verificaciones,
PuntosVerificacion, Evaluaciones, PlanesMantenimiento, Mantenimientos) —
si el cliente no las trae, se ignoran sin error, el MVP (Equipos/
Calibraciones) sigue funcionando igual. La regla de Evaluaciones de riesgo
ILAC reproduce EXACTAMENTE la regla real de routers/ilac.py (verificado en
el código): recalcula el intervalo sugerido con utils.calculos.
calcular_intervalo_inicial, y si no se declara un intervalo_adoptado_meses
explícito, lo deriva del intervalo real entre la 1ª y 2ª calibración del
historial del cliente (criterio de negocio de Edison, no inventado).
Verificado con generar_excel_prueba_fase5.py en modo --dry-run sin BD
(10/10 Crítica, 1/1 Alta, 2/2 Media exactos) — el modo --ejecutar de estas
6 hojas nuevas todavía NO se ha probado contra Postgres real (mismo
pendiente que tuvo el MVP hasta que se corrió el runbook de §9).

Uso:
    python importar_excel.py plantilla.xlsx
        -> dry-run, produce Registro_Desviaciones_<timestamp>.xlsx

    python importar_excel.py plantilla.xlsx --database-url postgresql://...
        -> dry-run + valida también duplicados contra esa BD (Capa 2/3)

    python importar_excel.py plantilla.xlsx --database-url postgresql://... \
        --ejecutar --registro-resueltas Registro_Desviaciones_resuelto.xlsx
        -> carga real, solo si no quedan Críticas y todas las Altas están
          resueltas/aceptadas en el registro suministrado
"""
import argparse
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl

# ---------------------------------------------------------------------------
# Esquema esperado - ver docs/migracion/PLAN_IMPORTACION_EXCEL.md §3
# ---------------------------------------------------------------------------

HOJAS_REQUERIDAS = ["Equipos", "Magnitudes", "Calibraciones", "PuntosCalibracion"]

# Fase 5 (docs/migracion/PLAN_FASE5_EXTENSIONES.md) - opcionales: si el
# cliente no las trae, leer_filas() devuelve listas vacías sin error, no
# bloquean la migración del MVP (Equipos/Calibraciones).
HOJAS_OPCIONALES = ["PlanesVerificacion", "Verificaciones", "PuntosVerificacion",
                     "Evaluaciones", "PlanesMantenimiento", "Mantenimientos"]

# Los 14 factores de riesgo ILAC-G24 - mismo orden/nombres que EvaluacionRiesgo
# en models.py y que el formulario real de routers/ilac.py.
FACTORES_ILAC = ["f_incertidumbre", "f_tipo", "f_riesgo_emp", "f_fabricante",
                  "f_deriva", "f_uso", "f_ambiental", "f_magnitud", "f_similares",
                  "f_comparaciones", "f_verificaciones", "f_transporte",
                  "f_personal", "f_legal"]

COLUMNAS = {
    "Equipos": {
        "obligatorias": ["codigo", "nombre"],
        "opcionales": ["descripcion", "marca", "modelo", "numero_serie",
                        "numero_inventario", "fecha_adquisicion", "costo",
                        "area", "ubicacion", "responsable", "estado"],
    },
    "Magnitudes": {
        "obligatorias": ["codigo_equipo", "nombre_magnitud"],
        "opcionales": ["simbolo", "unidad", "rango_min", "rango_max",
                        "resolucion", "emp_texto", "emp_valor", "emp_unidad",
                        "clase_exactitud", "tipo_instrumento"],
    },
    "Calibraciones": {
        "obligatorias": ["id_temporal", "codigo_equipo", "nombre_magnitud",
                          "fecha_calibracion"],
        "opcionales": ["numero_certificado", "laboratorio",
                        "acreditacion_laboratorio", "proxima_calibracion",
                        "patrones_utilizados", "metodo_calibracion",
                        "temperatura_ambiente", "humedad_relativa",
                        "trazabilidad", "observaciones", "costo", "resultado"],
    },
    "PuntosCalibracion": {
        "obligatorias": ["id_temporal_calibracion", "numero_punto",
                          "valor_patron", "valor_indicado"],
        "opcionales": ["incertidumbre", "observacion"],
    },
    # -- Fase 5 - docs/migracion/PLAN_FASE5_EXTENSIONES.md --
    "PlanesVerificacion": {
        "obligatorias": ["codigo_equipo", "nombre_magnitud"],
        "opcionales": ["frecuencia_meses", "procedimiento", "patron_referencia",
                        "umbral_alerta_pct", "umbral_fuera_pct", "activo",
                        "justificacion_no_aplica"],
    },
    "Verificaciones": {
        "obligatorias": ["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha"],
        "opcionales": ["proxima_verificacion", "tipo", "realizada_por", "patron_usado",
                        "resultado", "accion_tomada", "observaciones", "max_desviacion_pct"],
    },
    "PuntosVerificacion": {
        "obligatorias": ["id_temporal_verificacion", "numero_punto", "valor_patron"],
        "opcionales": ["valor_indicado", "tolerancia_inf", "tolerancia_sup", "observacion"],
    },
    "Evaluaciones": {
        "obligatorias": ["codigo_equipo", "nombre_magnitud"],
        "opcionales": FACTORES_ILAC + ["intervalo_fabricante_meses", "intervalo_adoptado_meses",
                        "justificacion", "justificacion_exceso", "evaluado_por"],
    },
    "PlanesMantenimiento": {
        "obligatorias": ["codigo_equipo"],
        "opcionales": ["frecuencia_meses", "tipo", "descripcion", "responsable", "activo"],
    },
    "Mantenimientos": {
        "obligatorias": ["codigo_equipo", "tipo", "origen", "titulo"],
        "opcionales": ["descripcion", "responsable_interno", "empresa_externa",
                        "tecnico_externo", "orden_trabajo", "fecha_programada",
                        "fecha_inicio", "fecha_fin", "estado", "falla_encontrada",
                        "trabajo_realizado", "repuestos_utilizados", "costo",
                        "requiere_calibracion", "afecta_medicion",
                        "observaciones_metrologicas"],
    },
}

ESTADOS_VALIDOS = {"operativo", "en_espera_calibracion", "fuera_de_uso", "dado_de_baja"}

CAMPOS_NUMERICOS = {
    "Equipos": {"costo"},
    "Magnitudes": {"rango_min", "rango_max", "emp_valor"},
    "Calibraciones": {"temperatura_ambiente", "humedad_relativa", "costo"},
    "PuntosCalibracion": {"numero_punto", "valor_patron", "valor_indicado", "incertidumbre"},
    "PlanesVerificacion": {"frecuencia_meses", "umbral_alerta_pct", "umbral_fuera_pct"},
    "Verificaciones": {"max_desviacion_pct"},
    "PuntosVerificacion": {"numero_punto", "valor_patron", "valor_indicado",
                            "tolerancia_inf", "tolerancia_sup"},
    "Evaluaciones": set(FACTORES_ILAC) | {"intervalo_fabricante_meses", "intervalo_adoptado_meses"},
    "PlanesMantenimiento": {"frecuencia_meses"},
    "Mantenimientos": {"costo"},
}
CAMPOS_FECHA = {
    "Equipos": {"fecha_adquisicion"},
    "Calibraciones": {"fecha_calibracion", "proxima_calibracion"},
    "Verificaciones": {"fecha", "proxima_verificacion"},
    "Mantenimientos": {"fecha_programada", "fecha_inicio", "fecha_fin"},
}


# ---------------------------------------------------------------------------
# Modelo de una desviación - docs/migracion/GUIA_VALIDACION_Y_DESVIACIONES.md §4
# ---------------------------------------------------------------------------

CRITICA, ALTA, MEDIA, BAJA = "Critica", "Alta", "Media", "Baja"


@dataclass
class Desviacion:
    hoja: str
    fila: int
    campo: str
    severidad: str
    regla_violada: str
    valor_encontrado: str
    valor_esperado_o_conflicto: str
    clave: str = ""  # clave de negocio estable - permite reconocer la misma
                      # desviación entre corridas aunque cambien los números
                      # de fila (ver GUIA_VALIDACION_Y_DESVIACIONES.md §5)


@dataclass
class ResultadoLectura:
    """Filas ya parseadas de cada hoja, listas para validar."""
    equipos: list = field(default_factory=list)
    magnitudes: list = field(default_factory=list)
    calibraciones: list = field(default_factory=list)
    puntos: list = field(default_factory=list)
    # -- Fase 5 - docs/migracion/PLAN_FASE5_EXTENSIONES.md --
    planes_verificacion: list = field(default_factory=list)
    verificaciones: list = field(default_factory=list)
    puntos_verificacion: list = field(default_factory=list)
    evaluaciones: list = field(default_factory=list)
    planes_mantenimiento: list = field(default_factory=list)
    mantenimientos: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Nivel 1 - estructural
# ---------------------------------------------------------------------------

def leer_libro(ruta, desviaciones):
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        print(f"ERROR: no se pudo abrir '{ruta}': {e}")
        sys.exit(2)

    for hoja in HOJAS_REQUERIDAS:
        if hoja not in wb.sheetnames:
            desviaciones.append(Desviacion(
                hoja=hoja, fila=0, campo="(hoja completa)", severidad=CRITICA,
                regla_violada="nivel1-estructural-hoja-faltante",
                valor_encontrado="(no existe)",
                valor_esperado_o_conflicto=f"Hoja obligatoria '{hoja}' según PLAN_IMPORTACION_EXCEL.md §3",
                clave=f"hoja:{hoja}",
            ))
    return wb


def leer_filas(wb, hoja, desviaciones):
    """Devuelve una lista de dicts {columna: valor}, con la fila real de
    Excel guardada en '_fila'. Valida cabeceras obligatorias (nivel 1)."""
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    filas_iter = ws.iter_rows(values_only=False)
    try:
        fila_header = next(filas_iter)
    except StopIteration:
        desviaciones.append(Desviacion(
            hoja=hoja, fila=1, campo="(cabecera)", severidad=CRITICA,
            regla_violada="nivel1-estructural-hoja-vacia",
            valor_encontrado="(hoja sin filas)",
            valor_esperado_o_conflicto="Al menos la fila de cabecera",
            clave=f"hoja:{hoja}",
        ))
        return []

    encabezados = [(c.value or "").strip() if isinstance(c.value, str) else c.value
                   for c in fila_header]
    col_idx = {nombre: i for i, nombre in enumerate(encabezados) if nombre}

    faltantes = [c for c in COLUMNAS[hoja]["obligatorias"] if c not in col_idx]
    for campo in faltantes:
        desviaciones.append(Desviacion(
            hoja=hoja, fila=1, campo=campo, severidad=CRITICA,
            regla_violada="nivel1-estructural-columna-faltante",
            valor_encontrado="(columna no presente)",
            valor_esperado_o_conflicto=f"Columna obligatoria '{campo}'",
            clave=f"hoja:{hoja}:col:{campo}",
        ))
    if faltantes:
        return []  # sin columnas obligatorias no tiene sentido seguir leyendo esta hoja

    todas_las_columnas = COLUMNAS[hoja]["obligatorias"] + COLUMNAS[hoja]["opcionales"]
    filas = []
    for n_fila, fila in enumerate(filas_iter, start=2):
        dato = {}
        vacia = True
        for campo in todas_las_columnas:
            if campo not in col_idx:
                dato[campo] = None
                continue
            celda = fila[col_idx[campo]] if col_idx[campo] < len(fila) else None
            valor = celda.value if celda is not None else None
            dato[campo] = valor
            if valor not in (None, ""):
                vacia = False
        if vacia:
            continue  # fila en blanco al final de la hoja, se ignora silenciosamente
        dato["_fila"] = n_fila
        filas.append(dato)
    return filas


# ---------------------------------------------------------------------------
# Nivel 1b - obligatorios vacíos por fila
# ---------------------------------------------------------------------------

def validar_obligatorios(hoja, filas, desviaciones):
    for f in filas:
        for campo in COLUMNAS[hoja]["obligatorias"]:
            if f.get(campo) in (None, ""):
                desviaciones.append(Desviacion(
                    hoja=hoja, fila=f["_fila"], campo=campo, severidad=CRITICA,
                    regla_violada="nivel1-estructural-valor-obligatorio-vacio",
                    valor_encontrado="(vacío)",
                    valor_esperado_o_conflicto=f"'{campo}' es obligatorio",
                    clave=f"{hoja}:{f['_fila']}:{campo}",
                ))


# ---------------------------------------------------------------------------
# Nivel 2 - tipo / formato
# ---------------------------------------------------------------------------

def _parsear_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parsear_bool(valor, default=None):
    """Parseo tolerante para columnas booleanas (activo, requiere_calibracion...).
    No genera desviación si no reconoce el valor - simplemente cae al default,
    a propósito para no sumar una cuarta capa de validación a un campo que no
    afecta integridad referencial ni cálculo metrológico."""
    if valor in (None, ""):
        return default
    if isinstance(valor, bool):
        return valor
    s = str(valor).strip().lower()
    if s in ("true", "1", "si", "sí", "x", "verdadero"):
        return True
    if s in ("false", "0", "no", "falso"):
        return False
    return default


def _meses_entre(f1, f2):
    """Meses completos entre dos fechas (fechas de calibración, para derivar
    el intervalo adoptado real del historial - ver PLAN_FASE5_EXTENSIONES.md
    §1.3). Convención: días / 30.44 (promedio real de días por mes), redondeado
    al entero más cercano - documentado a propósito, no es un cálculo oculto."""
    return round((f2 - f1).days / 30.44)


def validar_tipos(hoja, filas, desviaciones):
    for f in filas:
        for campo in CAMPOS_FECHA.get(hoja, set()):
            valor = f.get(campo)
            if valor in (None, ""):
                continue
            parsed = _parsear_fecha(valor)
            if parsed is None:
                sev = CRITICA if campo in COLUMNAS[hoja]["obligatorias"] else MEDIA
                desviaciones.append(Desviacion(
                    hoja=hoja, fila=f["_fila"], campo=campo, severidad=sev,
                    regla_violada="nivel2-tipo-fecha-no-parseable",
                    valor_encontrado=str(valor),
                    valor_esperado_o_conflicto="Fecha en formato AAAA-MM-DD o DD/MM/AAAA",
                    clave=f"{hoja}:{f['_fila']}:{campo}",
                ))
                f[campo] = None
            else:
                f[campo] = parsed
                hoy = date.today()
                if campo == "fecha_calibracion" and parsed > hoy:
                    desviaciones.append(Desviacion(
                        hoja=hoja, fila=f["_fila"], campo=campo, severidad=BAJA,
                        regla_violada="nivel2-fecha-futura-sospechosa",
                        valor_encontrado=str(parsed),
                        valor_esperado_o_conflicto="fecha_calibracion no debería ser futura para historial ya realizado",
                        clave=f"{hoja}:{f['_fila']}:{campo}",
                    ))
                if campo == "fecha_calibracion" and parsed.year < 1990:
                    desviaciones.append(Desviacion(
                        hoja=hoja, fila=f["_fila"], campo=campo, severidad=BAJA,
                        regla_violada="nivel2-fecha-muy-antigua-sospechosa",
                        valor_encontrado=str(parsed),
                        valor_esperado_o_conflicto="Confirmar que la fecha es correcta (año < 1990)",
                        clave=f"{hoja}:{f['_fila']}:{campo}",
                    ))

        for campo in CAMPOS_NUMERICOS.get(hoja, set()):
            valor = f.get(campo)
            if valor in (None, ""):
                if campo == "emp_valor" and hoja == "Magnitudes":
                    desviaciones.append(Desviacion(
                        hoja=hoja, fila=f["_fila"], campo=campo, severidad=MEDIA,
                        regla_violada="nivel2-emp-valor-ausente",
                        valor_encontrado="(vacío)",
                        valor_esperado_o_conflicto="Sin EMP, el semáforo de conformidad no se puede calcular para esta magnitud",
                        clave=f"{hoja}:{f['_fila']}:{campo}",
                    ))
                continue
            try:
                f[campo] = float(valor)
            except (TypeError, ValueError):
                sev = CRITICA if campo in COLUMNAS[hoja]["obligatorias"] else MEDIA
                desviaciones.append(Desviacion(
                    hoja=hoja, fila=f["_fila"], campo=campo, severidad=sev,
                    regla_violada="nivel2-tipo-numero-no-parseable",
                    valor_encontrado=str(valor),
                    valor_esperado_o_conflicto="Valor numérico",
                    clave=f"{hoja}:{f['_fila']}:{campo}",
                ))
                f[campo] = None

        if hoja == "Equipos":
            estado = f.get("estado")
            if estado and estado not in ESTADOS_VALIDOS:
                desviaciones.append(Desviacion(
                    hoja=hoja, fila=f["_fila"], campo="estado", severidad=MEDIA,
                    regla_violada="nivel2-estado-no-valido",
                    valor_encontrado=str(estado),
                    valor_esperado_o_conflicto=f"Uno de: {', '.join(sorted(ESTADOS_VALIDOS))}",
                    clave=f"{hoja}:{f['_fila']}:estado",
                ))


# ---------------------------------------------------------------------------
# Capa 1 de duplicados - dentro del propio Excel
# ---------------------------------------------------------------------------

def validar_duplicados_archivo(datos: ResultadoLectura, desviaciones):
    vistos = {}
    for f in datos.equipos:
        codigo = f.get("codigo")
        if not codigo:
            continue
        if codigo in vistos:
            desviaciones.append(Desviacion(
                hoja="Equipos", fila=f["_fila"], campo="codigo", severidad=CRITICA,
                regla_violada="capa1-duplicado-dentro-del-archivo",
                valor_encontrado=str(codigo),
                valor_esperado_o_conflicto=f"Ya aparece en la fila {vistos[codigo]} de esta misma hoja",
                clave=f"equipo:{codigo}",
            ))
        else:
            vistos[codigo] = f["_fila"]

    vistos_cal = {}
    for f in datos.calibraciones:
        idt = f.get("id_temporal")
        if not idt:
            continue
        if idt in vistos_cal:
            desviaciones.append(Desviacion(
                hoja="Calibraciones", fila=f["_fila"], campo="id_temporal", severidad=CRITICA,
                regla_violada="capa1-duplicado-dentro-del-archivo",
                valor_encontrado=str(idt),
                valor_esperado_o_conflicto=f"Ya aparece en la fila {vistos_cal[idt]} de esta misma hoja",
                clave=f"calibracion:{idt}",
            ))
        else:
            vistos_cal[idt] = f["_fila"]

    # -- Fase 5 --
    vistos_ver = {}
    for f in datos.verificaciones:
        idt = f.get("id_temporal")
        if not idt:
            continue
        if idt in vistos_ver:
            desviaciones.append(Desviacion(
                hoja="Verificaciones", fila=f["_fila"], campo="id_temporal", severidad=CRITICA,
                regla_violada="capa1-duplicado-dentro-del-archivo",
                valor_encontrado=str(idt),
                valor_esperado_o_conflicto=f"Ya aparece en la fila {vistos_ver[idt]} de esta misma hoja",
                clave=f"verificacion:{idt}",
            ))
        else:
            vistos_ver[idt] = f["_fila"]

    vistos_plan = {}
    for f in datos.planes_verificacion:
        clave = (f.get("codigo_equipo"), f.get("nombre_magnitud"))
        if not clave[0] or not clave[1]:
            continue
        if clave in vistos_plan:
            desviaciones.append(Desviacion(
                hoja="PlanesVerificacion", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="capa1-duplicado-dentro-del-archivo",
                valor_encontrado=f"{clave[0]} / {clave[1]}",
                valor_esperado_o_conflicto=f"Ya aparece en la fila {vistos_plan[clave]} de esta misma hoja",
                clave=f"plan_verificacion:{clave[0]}|{clave[1]}",
            ))
        else:
            vistos_plan[clave] = f["_fila"]

    vistos_eval = {}
    for f in datos.evaluaciones:
        clave = (f.get("codigo_equipo"), f.get("nombre_magnitud"))
        if not clave[0] or not clave[1]:
            continue
        if clave in vistos_eval:
            desviaciones.append(Desviacion(
                hoja="Evaluaciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="capa1-duplicado-dentro-del-archivo",
                valor_encontrado=f"{clave[0]} / {clave[1]}",
                valor_esperado_o_conflicto=f"Ya aparece en la fila {vistos_eval[clave]} de esta misma hoja "
                                            "(EvaluacionRiesgo es única por magnitud en la base de datos)",
                clave=f"evaluacion:{clave[0]}|{clave[1]}",
            ))
        else:
            vistos_eval[clave] = f["_fila"]


# ---------------------------------------------------------------------------
# Nivel 3 - referencial
# ---------------------------------------------------------------------------

def validar_referencial(datos: ResultadoLectura, desviaciones):
    codigos_equipo = {f["codigo"] for f in datos.equipos if f.get("codigo")}
    claves_magnitud = {(f["codigo_equipo"], f["nombre_magnitud"])
                        for f in datos.magnitudes
                        if f.get("codigo_equipo") and f.get("nombre_magnitud")}
    claves_calibracion = {f["id_temporal"] for f in datos.calibraciones if f.get("id_temporal")}

    for f in datos.magnitudes:
        ce = f.get("codigo_equipo")
        if ce and ce not in codigos_equipo:
            desviaciones.append(Desviacion(
                hoja="Magnitudes", fila=f["_fila"], campo="codigo_equipo", severidad=CRITICA,
                regla_violada="nivel3-referencial-equipo-inexistente",
                valor_encontrado=str(ce),
                valor_esperado_o_conflicto="Un código presente en la hoja Equipos",
                clave=f"magnitud:{ce}|{f.get('nombre_magnitud')}",
            ))

    for f in datos.calibraciones:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if ce and nm and (ce, nm) not in claves_magnitud:
            desviaciones.append(Desviacion(
                hoja="Calibraciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-magnitud-inexistente",
                valor_encontrado=f"{ce} / {nm}",
                valor_esperado_o_conflicto="Un par (codigo_equipo, nombre_magnitud) presente en la hoja Magnitudes",
                clave=f"calibracion:{f.get('id_temporal')}",
            ))

    for f in datos.puntos:
        idt = f.get("id_temporal_calibracion")
        if idt and idt not in claves_calibracion:
            n_punto = f.get("numero_punto")
            n_punto_str = str(int(n_punto)) if isinstance(n_punto, float) else str(n_punto)
            desviaciones.append(Desviacion(
                hoja="PuntosCalibracion", fila=f["_fila"], campo="id_temporal_calibracion",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-calibracion-inexistente",
                valor_encontrado=str(idt),
                valor_esperado_o_conflicto="Un id_temporal presente en la hoja Calibraciones",
                clave=f"punto:{idt}:{n_punto_str}",
            ))

    # -- Fase 5 - las hojas de Equipos/Magnitudes ya se validaron arriba;
    # aquí solo lo que es nuevo de docs/migracion/PLAN_FASE5_EXTENSIONES.md --
    claves_verificacion = {f["id_temporal"] for f in datos.verificaciones if f.get("id_temporal")}

    for f in datos.planes_verificacion:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if ce and nm and (ce, nm) not in claves_magnitud:
            desviaciones.append(Desviacion(
                hoja="PlanesVerificacion", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-magnitud-inexistente",
                valor_encontrado=f"{ce} / {nm}",
                valor_esperado_o_conflicto="Un par (codigo_equipo, nombre_magnitud) presente en la hoja Magnitudes",
                clave=f"plan_verificacion:{ce}|{nm}",
            ))

    for f in datos.verificaciones:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if ce and nm and (ce, nm) not in claves_magnitud:
            desviaciones.append(Desviacion(
                hoja="Verificaciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-magnitud-inexistente",
                valor_encontrado=f"{ce} / {nm}",
                valor_esperado_o_conflicto="Un par (codigo_equipo, nombre_magnitud) presente en la hoja Magnitudes",
                clave=f"verificacion:{f.get('id_temporal')}",
            ))
        # La existencia del PLAN (PlanesVerificacion o ya en BD) se valida
        # aparte en validar_verificaciones_plan_bd() porque puede depender
        # de la base de datos, no solo del archivo.

    for f in datos.puntos_verificacion:
        idt = f.get("id_temporal_verificacion")
        if idt and idt not in claves_verificacion:
            n_punto = f.get("numero_punto")
            n_punto_str = str(int(n_punto)) if isinstance(n_punto, float) else str(n_punto)
            desviaciones.append(Desviacion(
                hoja="PuntosVerificacion", fila=f["_fila"], campo="id_temporal_verificacion",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-verificacion-inexistente",
                valor_encontrado=str(idt),
                valor_esperado_o_conflicto="Un id_temporal presente en la hoja Verificaciones",
                clave=f"punto_verificacion:{idt}:{n_punto_str}",
            ))

    for f in datos.evaluaciones:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if ce and nm and (ce, nm) not in claves_magnitud:
            desviaciones.append(Desviacion(
                hoja="Evaluaciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-magnitud-inexistente",
                valor_encontrado=f"{ce} / {nm}",
                valor_esperado_o_conflicto="Un par (codigo_equipo, nombre_magnitud) presente en la hoja Magnitudes",
                clave=f"evaluacion:{ce}|{nm}",
            ))

    for f in datos.planes_mantenimiento:
        ce = f.get("codigo_equipo")
        if ce and ce not in codigos_equipo:
            desviaciones.append(Desviacion(
                hoja="PlanesMantenimiento", fila=f["_fila"], campo="codigo_equipo", severidad=CRITICA,
                regla_violada="nivel3-referencial-equipo-inexistente",
                valor_encontrado=str(ce),
                valor_esperado_o_conflicto="Un código presente en la hoja Equipos",
                clave=f"plan_mantenimiento:{ce}|{f.get('tipo') or 'preventivo'}",
            ))

    for f in datos.mantenimientos:
        ce = f.get("codigo_equipo")
        if ce and ce not in codigos_equipo:
            desviaciones.append(Desviacion(
                hoja="Mantenimientos", fila=f["_fila"], campo="codigo_equipo", severidad=CRITICA,
                regla_violada="nivel3-referencial-equipo-inexistente",
                valor_encontrado=str(ce),
                valor_esperado_o_conflicto="Un código presente en la hoja Equipos",
                clave=f"mantenimiento:{ce}|{f.get('titulo')}|{f['_fila']}",
            ))


# ---------------------------------------------------------------------------
# Nivel 4 - regla de negocio (reutiliza utils.calculos, no reimplementa)
# ---------------------------------------------------------------------------

def validar_regla_negocio(datos: ResultadoLectura, desviaciones):
    try:
        from utils.calculos import calcular_semaforo
    except Exception as e:
        print(f"AVISO: no se pudo importar utils.calculos ({e}) - se omite el nivel 4 "
              "(comparación de semáforo). Corre este script desde la carpeta del proyecto.")
        return

    emp_por_magnitud = {(f["codigo_equipo"], f["nombre_magnitud"]): f.get("emp_valor")
                         for f in datos.magnitudes}
    puntos_por_calibracion = {}
    for p in datos.puntos:
        puntos_por_calibracion.setdefault(p.get("id_temporal_calibracion"), []).append(p)

    for cal in datos.calibraciones:
        resultado_declarado = cal.get("resultado")
        if not resultado_declarado:
            continue  # nada que comparar
        emp = emp_por_magnitud.get((cal.get("codigo_equipo"), cal.get("nombre_magnitud")))
        if emp is None:
            continue  # ya se avisó en nivel 2 que falta el EMP
        puntos = puntos_por_calibracion.get(cal.get("id_temporal"), [])
        if not puntos:
            continue
        conformes = []
        for p in puntos:
            vp, vi = p.get("valor_patron"), p.get("valor_indicado")
            if vp is None or vi is None:
                continue
            error = vi - vp
            ok = calcular_semaforo(error, p.get("incertidumbre"), emp)
            if ok is not None:
                conformes.append(ok)
        if not conformes:
            continue
        recalculado = "aprobado" if all(conformes) else "rechazado"
        declarado_norm = str(resultado_declarado).strip().lower()
        if declarado_norm in ("aprobado", "rechazado") and declarado_norm != recalculado:
            desviaciones.append(Desviacion(
                hoja="Calibraciones", fila=cal["_fila"], campo="resultado", severidad=ALTA,
                regla_violada="nivel4-semaforo-no-coincide",
                valor_encontrado=str(resultado_declarado),
                valor_esperado_o_conflicto=f"Recalculado a partir de los puntos: '{recalculado}'",
                clave=f"calibracion:{cal.get('id_temporal')}",
            ))


# ---------------------------------------------------------------------------
# Fase 5 - regla propia de PlanesVerificacion (no depende de la BD)
# docs/migracion/PLAN_FASE5_EXTENSIONES.md §2.2 - severidad Media confirmada
# con Edison (12-ago-2026): se importa con advertencia, no bloquea.
# ---------------------------------------------------------------------------

def validar_plan_no_aplica(datos: ResultadoLectura, desviaciones):
    for f in datos.planes_verificacion:
        frecuencia = f.get("frecuencia_meses")
        justificacion = f.get("justificacion_no_aplica")
        if frecuencia in (None, "") and not (justificacion and str(justificacion).strip()):
            desviaciones.append(Desviacion(
                hoja="PlanesVerificacion", fila=f["_fila"], campo="justificacion_no_aplica",
                severidad=MEDIA,
                regla_violada="nivel2-plan-sin-frecuencia-ni-justificacion",
                valor_encontrado="(ambos vacíos)",
                valor_esperado_o_conflicto="Si el cliente no hace verificación intermedia para "
                                            "esta magnitud, documentar por qué en 'justificacion_no_aplica'",
                clave=f"plan_verificacion:{f.get('codigo_equipo')}|{f.get('nombre_magnitud')}",
            ))


# ---------------------------------------------------------------------------
# Capas 2 y 3 - duplicados contra la base de datos (opcional, requiere BD)
# ---------------------------------------------------------------------------

def validar_duplicados_bd(datos: ResultadoLectura, desviaciones, db):
    import models

    codigos_existentes = {c for (c,) in db.query(models.Equipo.codigo).all()}
    for f in datos.equipos:
        codigo = f.get("codigo")
        if codigo and codigo in codigos_existentes:
            desviaciones.append(Desviacion(
                hoja="Equipos", fila=f["_fila"], campo="codigo", severidad=BAJA,
                regla_violada="capa2-equipo-ya-existe-en-bd",
                valor_encontrado=str(codigo),
                valor_esperado_o_conflicto="Ya existe en la base de datos - no se modifica salvo --actualizar-existentes",
                clave=f"equipo:{codigo}",
            ))

    existentes = (
        db.query(models.Equipo.codigo, models.MagnitudEquipo.nombre,
                  models.Calibracion.fecha_calibracion, models.Calibracion.numero_certificado)
        .join(models.MagnitudEquipo, models.Calibracion.magnitud_id == models.MagnitudEquipo.id)
        .join(models.Equipo, models.Calibracion.equipo_id == models.Equipo.id)
        .all()
    )
    existentes_set = {(ec, mn, fc, nc) for ec, mn, fc, nc in existentes}

    for f in datos.calibraciones:
        clave_bd = (f.get("codigo_equipo"), f.get("nombre_magnitud"),
                    f.get("fecha_calibracion"), f.get("numero_certificado"))
        if clave_bd in existentes_set:
            desviaciones.append(Desviacion(
                hoja="Calibraciones", fila=f["_fila"], campo="(fila completa)", severidad=ALTA,
                regla_violada="capa3-calibracion-probable-duplicado-en-bd",
                valor_encontrado=f"{clave_bd[0]} / {clave_bd[1]} / {clave_bd[2]} / cert. {clave_bd[3]}",
                valor_esperado_o_conflicto="Ya existe una calibración con esta combinación en la base de datos",
                clave=f"calibracion:{f.get('id_temporal')}",
            ))

    # -- Fase 5 - EvaluacionRiesgo es única por magnitud en la BD
    # (magnitud_id, unique=True) - mismo tratamiento que equipos: informativa,
    # no se sobrescribe salvo --actualizar-existentes.
    magnitudes_ya_evaluadas = {
        (ec, mn) for ec, mn in
        db.query(models.Equipo.codigo, models.MagnitudEquipo.nombre)
        .select_from(models.EvaluacionRiesgo)
        .join(models.MagnitudEquipo, models.EvaluacionRiesgo.magnitud_id == models.MagnitudEquipo.id)
        .join(models.Equipo, models.MagnitudEquipo.equipo_id == models.Equipo.id)
        .all()
    }
    for f in datos.evaluaciones:
        clave = (f.get("codigo_equipo"), f.get("nombre_magnitud"))
        if clave[0] and clave[1] and clave in magnitudes_ya_evaluadas:
            desviaciones.append(Desviacion(
                hoja="Evaluaciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=BAJA,
                regla_violada="capa2-evaluacion-ya-existe-en-bd",
                valor_encontrado=f"{clave[0]} / {clave[1]}",
                valor_esperado_o_conflicto="Ya existe una evaluación de riesgo para esta magnitud - "
                                            "no se modifica salvo --actualizar-existentes",
                clave=f"evaluacion:{clave[0]}|{clave[1]}",
            ))


# ---------------------------------------------------------------------------
# Fase 5 - existencia del plan de verificación (archivo Y/o BD)
# docs/migracion/PLAN_FASE5_EXTENSIONES.md §2.1: una Verificación necesita un
# PlanVerificacion existente (plan_id es obligatorio en el modelo real). El
# plan puede venir en la propia hoja PlanesVerificacion de este archivo, o ya
# existir en la base de datos de una carga anterior - por eso esta función
# corre con o sin --database-url, igual que validar_duplicados_bd.
# ---------------------------------------------------------------------------

def validar_verificaciones_plan_bd(datos: ResultadoLectura, desviaciones, db=None):
    planes_disponibles = {
        (f["codigo_equipo"], f["nombre_magnitud"]) for f in datos.planes_verificacion
        if f.get("codigo_equipo") and f.get("nombre_magnitud")
    }
    if db is not None:
        import models
        filas_bd = (
            db.query(models.Equipo.codigo, models.MagnitudEquipo.nombre)
            .select_from(models.PlanVerificacion)
            .join(models.Equipo, models.PlanVerificacion.equipo_id == models.Equipo.id)
            .join(models.MagnitudEquipo, models.PlanVerificacion.magnitud_id == models.MagnitudEquipo.id)
            .all()
        )
        planes_disponibles |= set(filas_bd)

    for f in datos.verificaciones:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if ce and nm and (ce, nm) not in planes_disponibles:
            desviaciones.append(Desviacion(
                hoja="Verificaciones", fila=f["_fila"], campo="codigo_equipo/nombre_magnitud",
                severidad=CRITICA,
                regla_violada="nivel3-referencial-plan-verificacion-inexistente",
                valor_encontrado=f"{ce} / {nm}",
                valor_esperado_o_conflicto="Debe existir un plan para este equipo+magnitud en la "
                                            "hoja PlanesVerificacion de este archivo, o ya en la base de datos",
                clave=f"verificacion:{f.get('id_temporal')}",
            ))


# ---------------------------------------------------------------------------
# Fase 5 - Nivel 4 de Evaluaciones ILAC (reutiliza utils.calculos, no
# reimplementa - mismo principio que validar_regla_negocio). Reproduce
# EXACTAMENTE la regla real de routers/ilac.py líneas 66-73 (verificado en
# el código, no supuesto): sug = calcular_intervalo_inicial(factores, fab);
# si adoptado > sugerido sin justificacion_exceso, bloquea. Para historial
# migrado, el "adoptado" correcto es el intervalo que el cliente REALMENTE
# usó entre su 1ª y 2ª calibración (acordado con Edison, 12-ago-2026, ver
# PLAN_FASE5_EXTENSIONES.md §1.3) - no un valor inventado ni recalculado
# desde cero.
# ---------------------------------------------------------------------------

def validar_evaluaciones_ilac(datos: ResultadoLectura, desviaciones, db=None):
    try:
        from utils.calculos import calcular_intervalo_inicial
    except Exception as e:
        print(f"AVISO: no se pudo importar utils.calculos ({e}) - se omite la validación de ILAC.")
        return

    fechas_por_clave = {}
    for c in datos.calibraciones:
        clave = (c.get("codigo_equipo"), c.get("nombre_magnitud"))
        fc = c.get("fecha_calibracion")
        if clave[0] and clave[1] and fc:
            fechas_por_clave.setdefault(clave, []).append(fc)

    if db is not None:
        import models
        filas_bd = (
            db.query(models.Equipo.codigo, models.MagnitudEquipo.nombre,
                      models.Calibracion.fecha_calibracion)
            .join(models.MagnitudEquipo, models.Calibracion.magnitud_id == models.MagnitudEquipo.id)
            .join(models.Equipo, models.Calibracion.equipo_id == models.Equipo.id)
            .all()
        )
        for ec, mn, fc in filas_bd:
            if fc:
                fechas_por_clave.setdefault((ec, mn), []).append(fc)

    for f in datos.evaluaciones:
        ce, nm = f.get("codigo_equipo"), f.get("nombre_magnitud")
        if not ce or not nm:
            continue
        factores = [f.get(campo) if f.get(campo) is not None else 3 for campo in FACTORES_ILAC]
        fab = f.get("intervalo_fabricante_meses")
        sug = calcular_intervalo_inicial(factores, fab)

        ado_declarado = f.get("intervalo_adoptado_meses")
        if ado_declarado is not None:
            ado = int(ado_declarado)
            origen_ado = "declarado en la hoja Evaluaciones"
        else:
            fechas = sorted(set(fechas_por_clave.get((ce, nm), [])))
            if len(fechas) >= 2:
                ado = _meses_entre(fechas[0], fechas[1])
                origen_ado = f"derivado del historial real ({fechas[0]} -> {fechas[1]})"
            else:
                ado = sug
                origen_ado = "sin 2 calibraciones de historial disponibles, igual al sugerido"

        justificacion_exceso = f.get("justificacion_exceso")
        if ado > sug and not (justificacion_exceso and str(justificacion_exceso).strip()):
            desviaciones.append(Desviacion(
                hoja="Evaluaciones", fila=f["_fila"], campo="intervalo_adoptado_meses",
                severidad=ALTA,
                regla_violada="nivel4-ilac-adoptado-excede-sugerido-sin-justificacion",
                valor_encontrado=f"adoptado={ado} ({origen_ado})",
                valor_esperado_o_conflicto=f"sugerido={sug} meses (14 factores de riesgo) - "
                                            "falta 'justificacion_exceso'",
                clave=f"evaluacion:{ce}|{nm}",
            ))

        # Se guarda el resultado ya resuelto en la propia fila para que
        # ejecutar_carga() no tenga que volver a calcularlo (y para que no
        # pueda desviarse entre la validación y la escritura real).
        f["_intervalo_adoptado_resuelto"] = ado
        f["_intervalo_sugerido_resuelto"] = sug
        f["_puntuacion_resuelta"] = round(sum(factores) / len(factores), 2)


# ---------------------------------------------------------------------------
# Registro de Desviaciones - salida
# ---------------------------------------------------------------------------

def escribir_registro(desviaciones, ruta_salida, resueltas=None):
    resueltas = resueltas or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Desviaciones"
    encabezados = ["id_desviacion", "hoja", "fila", "campo", "severidad",
                   "regla_violada", "valor_encontrado", "valor_esperado_o_conflicto",
                   "clave", "fecha_deteccion", "estado", "decision", "decidido_por",
                   "fecha_cierre"]
    ws.append(encabezados)
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M")
    orden_severidad = {CRITICA: 0, ALTA: 1, MEDIA: 2, BAJA: 3}
    desviaciones_ordenadas = sorted(desviaciones, key=lambda d: (orden_severidad[d.severidad], d.hoja, d.fila))
    for i, d in enumerate(desviaciones_ordenadas, start=1):
        resuelta = resueltas.get(d.clave)
        estado = resuelta["estado"] if resuelta else "Abierta"
        decision = resuelta["decision"] if resuelta else ""
        decidido_por = resuelta["decidido_por"] if resuelta else ""
        fecha_cierre = resuelta["fecha_cierre"] if resuelta else ""
        ws.append([f"DEV-{i:03d}", d.hoja, d.fila, d.campo, d.severidad, d.regla_violada,
                   d.valor_encontrado, d.valor_esperado_o_conflicto, d.clave, ahora,
                   estado, decision, decidido_por, fecha_cierre])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    wb.save(ruta_salida)


def cargar_registro_resueltas(ruta):
    """Lee un Registro de Desviaciones ya trabajado (ver GUIA_VALIDACION_Y_
    DESVIACIONES.md §5) y devuelve {clave: {estado, decision, decidido_por, fecha_cierre}}
    para las filas marcadas como Resuelta o Aceptada.

    IMPORTANTE (13-ago-2026, hallazgo real del runbook de
    PLAN_FASE5_EXTENSIONES.md §7): 'Resuelta - corregida en origen' NO
    cuenta aquí como resuelta. Ese estado es una PROMESA de que la fila se
    quitó/corrigió en el Excel de origen y se volvió a generar la plantilla
    - si esa promesa se cumplió, la clave simplemente no debería volver a
    aparecer en una corrida nueva, y este archivo de --registro-resueltas
    ni se necesita para ella. Si la clave SÍ reaparece, es señal de que la
    corrección no se aplicó de verdad - dejarla pasar solo porque alguna
    vez alguien escribió 'corregida en origen' sería confiar ciegamente en
    una nota vieja, exactamente lo que este framework existe para evitar
    (ver GUIA_VALIDACION_Y_DESVIACIONES.md §3, advertencia sobre Capa 3).
    Solo 'Aceptada...' (decisión consciente de insertar el dato tal cual
    está) y 'Resuelta - corrección manual documentada' (dato corregido
    directamente en la plantilla, con justificación) cuentan como
    definitivamente resueltas para --ejecutar."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {}
    encabezados = filas[0]
    idx = {nombre: i for i, nombre in enumerate(encabezados)}
    resultado = {}
    for fila in filas[1:]:
        if not fila or fila[idx.get("clave", -1)] in (None, ""):
            continue
        estado = fila[idx["estado"]] or ""
        estado_norm = str(estado).lower()
        if "corregida en origen" in estado_norm:
            continue  # no cuenta - ver docstring: si reaparece, sigue bloqueando
        if not estado_norm.startswith(("resuelta", "aceptada")):
            continue
        resultado[fila[idx["clave"]]] = {
            "estado": estado,
            "decision": fila[idx.get("decision", -1)] if "decision" in idx else "",
            "decidido_por": fila[idx.get("decidido_por", -1)] if "decidido_por" in idx else "",
            "fecha_cierre": fila[idx.get("fecha_cierre", -1)] if "fecha_cierre" in idx else "",
        }
    return resultado


def imprimir_resumen(desviaciones):
    conteo = {CRITICA: 0, ALTA: 0, MEDIA: 0, BAJA: 0}
    for d in desviaciones:
        conteo[d.severidad] += 1
    print("\n-- Resumen de desviaciones --------------------------------")
    print(f"  Críticas: {conteo[CRITICA]}   (bloquean SIEMPRE la carga)")
    print(f"  Altas:    {conteo[ALTA]}   (bloquean hasta decisión con el cliente)")
    print(f"  Medias:   {conteo[MEDIA]}   (se importan con advertencia)")
    print(f"  Bajas:    {conteo[BAJA]}   (informativas)")
    print("------------------------------------------------------------\n")
    return conteo


# ---------------------------------------------------------------------------
# Ejecución real de la carga (requiere BD, requiere --ejecutar)
# ---------------------------------------------------------------------------

def ejecutar_carga(datos: ResultadoLectura, db, actualizar_existentes=False):
    """Inserta en el orden Equipo -> Magnitud -> Calibracion -> Punto,
    dentro de UNA sola transacción (mismo patrón que seed_demo_data.py:
    db.flush() para obtener el ID antes de crear los hijos). El caller
    (main()) ya garantizó que no quedan desviaciones Críticas y que todas
    las Altas están resueltas/aceptadas - esta función no vuelve a decidir
    nada, solo inserta.

    NOTA - limitación conocida, documentada a propósito en vez de fingir
    que no existe: si un equipo ya existía y ya tenía magnitudes con el
    mismo nombre, este MVP reutiliza esa magnitud existente en vez de
    duplicarla, pero no fusiona/actualiza sus campos (emp_valor, unidad,
    etc.) - eso queda para una iteración futura si un cliente real lo
    necesita (ver docs/migracion/PLAN_IMPORTACION_EXCEL.md §7)."""
    import models

    equipo_id_por_codigo = dict(
        (c, eid) for eid, c in db.query(models.Equipo.id, models.Equipo.codigo).all()
    )

    creados = {"equipos": 0, "magnitudes": 0, "calibraciones": 0, "puntos": 0,
               "planes_verificacion": 0, "verificaciones": 0, "puntos_verificacion": 0,
               "evaluaciones": 0, "planes_mantenimiento": 0, "mantenimientos": 0}
    omitidos_equipos_existentes = 0

    for f in datos.equipos:
        codigo = f["codigo"]
        if codigo in equipo_id_por_codigo:
            omitidos_equipos_existentes += 1
            if not actualizar_existentes:
                continue
        eq = models.Equipo(
            codigo=codigo, nombre=f["nombre"], descripcion=f.get("descripcion"),
            marca=f.get("marca"), modelo=f.get("modelo"), numero_serie=f.get("numero_serie"),
            numero_inventario=f.get("numero_inventario"), fecha_adquisicion=f.get("fecha_adquisicion"),
            costo=f.get("costo"), area=f.get("area"), ubicacion=f.get("ubicacion"),
            responsable=f.get("responsable"),
            estado=f.get("estado") if f.get("estado") in ESTADOS_VALIDOS else "en_espera_calibracion",
        )
        db.add(eq)
        db.flush()
        equipo_id_por_codigo[codigo] = eq.id
        creados["equipos"] += 1

    magnitud_id_por_clave = {}
    existentes_mag = (
        db.query(models.MagnitudEquipo.id, models.Equipo.codigo, models.MagnitudEquipo.nombre)
        .join(models.Equipo, models.MagnitudEquipo.equipo_id == models.Equipo.id)
        .all()
    )
    for mid, codigo_eq, nombre_mag in existentes_mag:
        magnitud_id_por_clave[(codigo_eq, nombre_mag)] = mid

    for f in datos.magnitudes:
        clave = (f["codigo_equipo"], f["nombre_magnitud"])
        if clave in magnitud_id_por_clave:
            continue  # reutilizar existente - ver limitación en el docstring
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        if equipo_id is None:
            continue  # no debería pasar: nivel 3 ya lo habría marcado Crítico
        mag = models.MagnitudEquipo(
            equipo_id=equipo_id, nombre=f["nombre_magnitud"], simbolo=f.get("simbolo"),
            unidad=f.get("unidad"), rango_min=f.get("rango_min"), rango_max=f.get("rango_max"),
            resolucion=f.get("resolucion"), emp_texto=f.get("emp_texto"),
            emp_valor=f.get("emp_valor"), emp_unidad=f.get("emp_unidad"),
            clase_exactitud=f.get("clase_exactitud"),
            tipo_instrumento=f.get("tipo_instrumento") or "continuo",
        )
        db.add(mag)
        db.flush()
        magnitud_id_por_clave[clave] = mag.id
        creados["magnitudes"] += 1

    calibracion_id_por_temporal = {}
    for f in datos.calibraciones:
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        magnitud_id = magnitud_id_por_clave.get((f["codigo_equipo"], f["nombre_magnitud"]))
        if equipo_id is None or magnitud_id is None:
            continue
        cal = models.Calibracion(
            magnitud_id=magnitud_id, equipo_id=equipo_id,
            numero_certificado=f.get("numero_certificado"), laboratorio=f.get("laboratorio"),
            acreditacion_laboratorio=f.get("acreditacion_laboratorio"),
            fecha_calibracion=f["fecha_calibracion"], proxima_calibracion=f.get("proxima_calibracion"),
            patrones_utilizados=f.get("patrones_utilizados"), metodo_calibracion=f.get("metodo_calibracion"),
            temperatura_ambiente=f.get("temperatura_ambiente"), humedad_relativa=f.get("humedad_relativa"),
            trazabilidad=f.get("trazabilidad"), observaciones=f.get("observaciones"),
            costo=f.get("costo"), resultado=f.get("resultado") or "pendiente",
        )
        db.add(cal)
        db.flush()
        calibracion_id_por_temporal[f["id_temporal"]] = cal.id
        creados["calibraciones"] += 1

    for f in datos.puntos:
        cal_id = calibracion_id_por_temporal.get(f["id_temporal_calibracion"])
        if cal_id is None:
            continue
        pto = models.PuntoCalibracion(
            calibracion_id=cal_id, numero_punto=int(f["numero_punto"]),
            valor_patron=f["valor_patron"], valor_indicado=f["valor_indicado"],
            incertidumbre=f.get("incertidumbre"), observacion=f.get("observacion"),
        )
        db.add(pto)
        creados["puntos"] += 1

    # -- Fase 5 - Planes de verificación --
    plan_id_por_clave = {}
    for pid, codigo_eq, nombre_mag in (
        db.query(models.PlanVerificacion.id, models.Equipo.codigo, models.MagnitudEquipo.nombre)
        .join(models.Equipo, models.PlanVerificacion.equipo_id == models.Equipo.id)
        .join(models.MagnitudEquipo, models.PlanVerificacion.magnitud_id == models.MagnitudEquipo.id)
        .all()
    ):
        plan_id_por_clave[(codigo_eq, nombre_mag)] = pid

    for f in datos.planes_verificacion:
        clave = (f["codigo_equipo"], f["nombre_magnitud"])
        if clave in plan_id_por_clave:
            continue  # ya existe - no se modifica (mismo criterio que equipos/magnitudes)
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        magnitud_id = magnitud_id_por_clave.get(clave)
        if equipo_id is None or magnitud_id is None:
            continue
        plan = models.PlanVerificacion(
            equipo_id=equipo_id, magnitud_id=magnitud_id,
            frecuencia_meses=int(f["frecuencia_meses"]) if f.get("frecuencia_meses") else None,
            procedimiento=f.get("procedimiento"), patron_referencia=f.get("patron_referencia"),
            umbral_alerta_pct=f.get("umbral_alerta_pct") if f.get("umbral_alerta_pct") is not None else 70.0,
            umbral_fuera_pct=f.get("umbral_fuera_pct") if f.get("umbral_fuera_pct") is not None else 100.0,
            activo=_parsear_bool(f.get("activo"), True),
            justificacion_no_aplica=f.get("justificacion_no_aplica"),
        )
        db.add(plan)
        db.flush()
        plan_id_por_clave[clave] = plan.id
        creados["planes_verificacion"] += 1

    # -- Fase 5 - Verificaciones --
    verificacion_id_por_temporal = {}
    for f in datos.verificaciones:
        clave = (f["codigo_equipo"], f["nombre_magnitud"])
        plan_id = plan_id_por_clave.get(clave)
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        magnitud_id = magnitud_id_por_clave.get(clave)
        if plan_id is None or equipo_id is None or magnitud_id is None:
            continue  # no debería pasar: validar_verificaciones_plan_bd ya lo habría marcado Crítico
        ver = models.VerificacionIntermedia(
            plan_id=plan_id, equipo_id=equipo_id, magnitud_id=magnitud_id,
            fecha=f["fecha"], proxima_verificacion=f.get("proxima_verificacion"),
            tipo=f.get("tipo") or "programada", realizada_por=f.get("realizada_por"),
            patron_usado=f.get("patron_usado"), resultado=f.get("resultado") or "pendiente",
            accion_tomada=f.get("accion_tomada"), observaciones=f.get("observaciones"),
            max_desviacion_pct=f.get("max_desviacion_pct"),
        )
        db.add(ver)
        db.flush()
        verificacion_id_por_temporal[f["id_temporal"]] = ver.id
        creados["verificaciones"] += 1

    # -- Fase 5 - Puntos de verificación --
    for f in datos.puntos_verificacion:
        ver_id = verificacion_id_por_temporal.get(f["id_temporal_verificacion"])
        if ver_id is None:
            continue
        pv = models.PuntoVerificacion(
            verificacion_id=ver_id, numero_punto=int(f["numero_punto"]),
            valor_patron=f["valor_patron"], valor_indicado=f.get("valor_indicado"),
            tolerancia_inf=f.get("tolerancia_inf"), tolerancia_sup=f.get("tolerancia_sup"),
            observacion=f.get("observacion"),
        )
        db.add(pv)
        creados["puntos_verificacion"] += 1

    # -- Fase 5 - Evaluaciones de riesgo ILAC --
    # intervalo_adoptado/sugerido/puntuacion ya vienen resueltos por
    # validar_evaluaciones_ilac() - nunca se recalculan aquí para que no
    # puedan desviarse entre lo que se validó y lo que se escribe.
    magnitudes_ya_evaluadas = {mid for (mid,) in db.query(models.EvaluacionRiesgo.magnitud_id).all()}
    for f in datos.evaluaciones:
        clave = (f["codigo_equipo"], f["nombre_magnitud"])
        magnitud_id = magnitud_id_por_clave.get(clave)
        if magnitud_id is None:
            continue
        if magnitud_id in magnitudes_ya_evaluadas and not actualizar_existentes:
            continue  # ya existe - no se modifica (mismo criterio que equipos)
        valores_factores = {campo: (f.get(campo) if f.get(campo) is not None else 3)
                             for campo in FACTORES_ILAC}
        ev = models.EvaluacionRiesgo(
            magnitud_id=magnitud_id, **valores_factores,
            intervalo_fabricante_meses=f.get("intervalo_fabricante_meses"),
            puntuacion_total=f.get("_puntuacion_resuelta"),
            intervalo_sugerido_meses=f.get("_intervalo_sugerido_resuelto"),
            intervalo_adoptado_meses=f.get("_intervalo_adoptado_resuelto"),
            justificacion=f.get("justificacion"), justificacion_exceso=f.get("justificacion_exceso"),
            evaluado_por=f.get("evaluado_por"),
        )
        db.add(ev)
        magnitudes_ya_evaluadas.add(magnitud_id)
        creados["evaluaciones"] += 1

    # -- Fase 5 - Planes de mantenimiento --
    planes_mtto_existentes = set(
        db.query(models.Equipo.codigo, models.PlanMantenimiento.tipo)
        .join(models.PlanMantenimiento, models.PlanMantenimiento.equipo_id == models.Equipo.id)
        .all()
    )
    for f in datos.planes_mantenimiento:
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        if equipo_id is None:
            continue
        tipo = f.get("tipo") or "preventivo"
        if (f["codigo_equipo"], tipo) in planes_mtto_existentes and not actualizar_existentes:
            continue
        pm = models.PlanMantenimiento(
            equipo_id=equipo_id,
            frecuencia_meses=int(f["frecuencia_meses"]) if f.get("frecuencia_meses") else 6,
            tipo=tipo, descripcion=f.get("descripcion"), responsable=f.get("responsable"),
            activo=_parsear_bool(f.get("activo"), True),
        )
        db.add(pm)
        planes_mtto_existentes.add((f["codigo_equipo"], tipo))
        creados["planes_mantenimiento"] += 1

    # -- Fase 5 - Mantenimientos --
    # Sin detección de duplicados (a diferencia de Calibraciones) - limitación
    # conocida y documentada a propósito: el modelo no define una clave
    # natural clara para "el mismo mantenimiento" (a diferencia del número de
    # certificado en Calibraciones). LO MISMO APLICA A Verificaciones/
    # PuntosVerificacion arriba: id_temporal es una clave solo dentro del
    # archivo, no se persiste en la BD, así que si el mismo archivo se
    # --ejecutar dos veces, las Verificaciones SÍ se duplican (a diferencia
    # de Equipos/Calibraciones/Evaluaciones/Planes, que sí detectan que ya
    # existen y no repiten). Si un cliente real necesita re-ejecutar de forma
    # segura, se agrega una Capa 2/3-equivalente para Verificaciones cuando
    # aparezca el caso concreto - ver PLAN_FASE5_EXTENSIONES.md.
    for f in datos.mantenimientos:
        equipo_id = equipo_id_por_codigo.get(f["codigo_equipo"])
        if equipo_id is None:
            continue
        mto = models.Mantenimiento(
            equipo_id=equipo_id, tipo=f["tipo"], origen=f["origen"], titulo=f["titulo"],
            descripcion=f.get("descripcion"), responsable_interno=f.get("responsable_interno"),
            empresa_externa=f.get("empresa_externa"), tecnico_externo=f.get("tecnico_externo"),
            orden_trabajo=f.get("orden_trabajo"), fecha_programada=f.get("fecha_programada"),
            fecha_inicio=f.get("fecha_inicio"), fecha_fin=f.get("fecha_fin"),
            estado=f.get("estado") or "programado",
            falla_encontrada=f.get("falla_encontrada"), trabajo_realizado=f.get("trabajo_realizado"),
            repuestos_utilizados=f.get("repuestos_utilizados"), costo=f.get("costo"),
            requiere_calibracion=_parsear_bool(f.get("requiere_calibracion"), False),
            afecta_medicion=_parsear_bool(f.get("afecta_medicion"), False),
            observaciones_metrologicas=f.get("observaciones_metrologicas"),
        )
        db.add(mto)
        creados["mantenimientos"] += 1

    creados["equipos_ya_existentes_omitidos"] = omitidos_equipos_existentes
    return creados


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("plantilla", help="Ruta al archivo .xlsx de la plantilla estándar (ya transformado por Power Query)")
    ap.add_argument("--database-url", default=os.getenv("MIGRACION_DATABASE_URL"),
                     help="URL de conexión (o variable de entorno MIGRACION_DATABASE_URL). "
                          "Sin esto, solo se valida el archivo, sin comparar contra la BD ni poder ejecutar.")
    ap.add_argument("--salida", default=None, help="Ruta del Registro de Desviaciones a generar")
    ap.add_argument("--registro-resueltas", default=None,
                     help="Registro de Desviaciones ya trabajado con el cliente (estado Resuelta/Aceptada)")
    ap.add_argument("--ejecutar", action="store_true", help="Escribe realmente en la base de datos (por defecto es solo dry-run)")
    ap.add_argument("--actualizar-existentes", action="store_true",
                     help="Si un equipo ya existe, actualizar sus campos en vez de omitirlo (por defecto NO se toca)")
    ap.add_argument("--confirmo-produccion", action="store_true",
                     help="Confirma a propósito que sabes que la URL no parece de staging")
    ap.add_argument("--sin-confirmacion", action="store_true",
                     help="Omite la confirmación interactiva 'SI' - SOLO para pruebas automatizadas, nunca para una migración real")
    ap.add_argument("--usuario-id", type=int, default=None,
                     help="ID del usuario (tabla usuarios) responsable de esta migración - queda "
                          "registrado en cada fila de registro_auditoria en vez de quedar en NULL. "
                          "Recomendado siempre que se cargue contra una base real de cliente.")
    args = ap.parse_args()

    desviaciones = []
    wb = leer_libro(args.plantilla, desviaciones)

    datos = ResultadoLectura()
    datos.equipos = leer_filas(wb, "Equipos", desviaciones)
    datos.magnitudes = leer_filas(wb, "Magnitudes", desviaciones)
    datos.calibraciones = leer_filas(wb, "Calibraciones", desviaciones)
    datos.puntos = leer_filas(wb, "PuntosCalibracion", desviaciones)
    # -- Fase 5 - hojas opcionales; leer_filas() devuelve [] sin error si el
    # cliente no las trae (ver HOJAS_OPCIONALES) --
    datos.planes_verificacion = leer_filas(wb, "PlanesVerificacion", desviaciones)
    datos.verificaciones = leer_filas(wb, "Verificaciones", desviaciones)
    datos.puntos_verificacion = leer_filas(wb, "PuntosVerificacion", desviaciones)
    datos.evaluaciones = leer_filas(wb, "Evaluaciones", desviaciones)
    datos.planes_mantenimiento = leer_filas(wb, "PlanesMantenimiento", desviaciones)
    datos.mantenimientos = leer_filas(wb, "Mantenimientos", desviaciones)

    for hoja, filas in (
        ("Equipos", datos.equipos), ("Magnitudes", datos.magnitudes),
        ("Calibraciones", datos.calibraciones), ("PuntosCalibracion", datos.puntos),
        ("PlanesVerificacion", datos.planes_verificacion), ("Verificaciones", datos.verificaciones),
        ("PuntosVerificacion", datos.puntos_verificacion), ("Evaluaciones", datos.evaluaciones),
        ("PlanesMantenimiento", datos.planes_mantenimiento), ("Mantenimientos", datos.mantenimientos),
    ):
        validar_obligatorios(hoja, filas, desviaciones)
        validar_tipos(hoja, filas, desviaciones)

    validar_duplicados_archivo(datos, desviaciones)
    validar_referencial(datos, desviaciones)
    validar_regla_negocio(datos, desviaciones)
    validar_plan_no_aplica(datos, desviaciones)  # Fase 5 - file-only

    db = None
    if args.database_url:
        try:
            from sqlalchemy import create_engine
            from sqlalchemy.orm import sessionmaker
            # Importar este módulo es lo que registra los listeners de
            # before_flush/after_flush que generan el rastro de auditoría
            # automático (utils/auditoria_trail.py). Si no se importa aquí,
            # el proceso de este script nunca los registra -- son eventos a
            # nivel de sqlalchemy.orm.Session que solo se activan cuando el
            # módulo se ha importado en algún punto del proceso. Se
            # descubrió que faltaba corriendo el runbook de §9 contra un
            # Postgres real: la carga escribía bien, pero no generaba NINGUNA
            # fila en registro_auditoria (confirmado con evidencia, no
            # supuesto). Ver docs/migracion/PLAN_IMPORTACION_EXCEL.md §9.
            import utils.auditoria_trail as auditoria_trail
            if args.usuario_id is not None:
                auditoria_trail.usuario_actual_id.set(args.usuario_id)
            engine = create_engine(args.database_url)
            Session = sessionmaker(bind=engine)
            db = Session()
            validar_duplicados_bd(datos, desviaciones, db)
        except Exception as e:
            print(f"AVISO: no se pudo conectar a la base de datos ({e}) - se omiten las "
                  "Capas 2/3 de duplicados. Revisa --database-url.")
            db = None
    else:
        print("AVISO: sin --database-url, no se validan duplicados contra la base de datos "
              "(Capa 2/3) ni se puede ejecutar la carga real. Esto es válido para que el "
              "cliente revise su archivo antes de enviarlo.")

    # Fase 5 - corren con o sin BD conectada (db puede ser None); si hay BD,
    # también miran el historial/planes ya existentes ahí, no solo el archivo.
    validar_verificaciones_plan_bd(datos, desviaciones, db)
    validar_evaluaciones_ilac(datos, desviaciones, db)

    resueltas = cargar_registro_resueltas(args.registro_resueltas) if args.registro_resueltas else {}

    ruta_salida = args.salida or f"Registro_Desviaciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    escribir_registro(desviaciones, ruta_salida, resueltas)
    conteo = imprimir_resumen(desviaciones)
    print(f"Registro de desviaciones escrito en: {ruta_salida}\n")

    if not args.ejecutar:
        if conteo[CRITICA] == 0 and conteo[ALTA] == 0:
            print("Sin desviaciones Críticas ni Altas pendientes - listo para --ejecutar.")
        else:
            print("Quedan desviaciones Críticas y/o Altas - resuélvelas con el cliente "
                  "(ver docs/migracion/GUIA_VALIDACION_Y_DESVIACIONES.md §5) antes de --ejecutar.")
        return

    # -- A partir de aquí, modo --ejecutar: validar TODO antes de escribir --
    if conteo[CRITICA] > 0:
        print(f"ABORTADO: hay {conteo[CRITICA]} desviación(es) Crítica(s) - no se puede ejecutar la carga.")
        sys.exit(1)

    altas_no_resueltas = [d for d in desviaciones if d.severidad == ALTA and d.clave not in resueltas]
    if altas_no_resueltas:
        print(f"ABORTADO: hay {len(altas_no_resueltas)} desviación(es) Alta(s) sin resolver/aceptar "
              f"en --registro-resueltas:")
        for d in altas_no_resueltas[:10]:
            print(f"  - {d.hoja} fila {d.fila} ({d.clave}): {d.regla_violada}")
        print("\nSi alguna de estas ya la marcaste 'Resuelta - corregida en origen' y sigue "
              "apareciendo, la corrección no llegó a este archivo (revisa que estés usando la "
              "plantilla ya corregida). Si en realidad el dato es correcto tal como está, "
              "cámbiala a 'Aceptada' en vez de 'Resuelta - corregida en origen' - ver "
              "docs/migracion/GUIA_VALIDACION_Y_DESVIACIONES.md §3.")
        sys.exit(1)

    if db is None:
        print("ABORTADO: --ejecutar requiere --database-url (o MIGRACION_DATABASE_URL).")
        sys.exit(1)

    if not args.confirmo_produccion and not any(
        x in args.database_url.lower() for x in ("test", "staging", "stg", "carga")
    ):
        print("ABORTADO: la URL de base de datos no parece de staging/prueba.\n"
              "Recomendado: correr primero contra una copia de staging.\n"
              "Si de verdad quieres cargar directo contra esta base de datos, "
              "vuelve a correr con --confirmo-produccion.")
        sys.exit(1)

    destino = args.database_url.split('@')[-1] if '@' in args.database_url else args.database_url
    print(f"Se van a insertar datos reales en: {destino}")
    print("¿Ya hiciste un backup con backup_db.py? Esta operación no se puede deshacer "
          "sola si algo sale mal después de confirmar.")
    if not args.sin_confirmacion:
        respuesta = input("Escribe SI (mayúsculas) para continuar: ")
        if respuesta.strip() != "SI":
            print("Cancelado por el usuario.")
            sys.exit(1)

    try:
        creados = ejecutar_carga(datos, db, actualizar_existentes=args.actualizar_existentes)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"ERROR durante la carga - se revirtió todo, no quedó nada a medias: {e}")
        sys.exit(1)

    print("\n-- Carga completada -----------------------------------")
    for k, v in creados.items():
        print(f"  {k}: {v}")
    print("---------------------------------------------------------")
    print("Recuerda archivar el Registro de Desviaciones final junto con la evidencia "
          "de esta migración (ver GUIA_VALIDACION_Y_DESVIACIONES.md §5, paso 7).")


if __name__ == "__main__":
    main()
