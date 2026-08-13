#!/usr/bin/env python3
"""
generar_excel_prueba_fase5_limpio.py — Excel SIN errores deliberados,
con las 6 hojas nuevas de la Fase 5, para probar el camino positivo
(--ejecutar real contra Postgres) — ver el runbook de
docs/migracion/PLAN_FASE5_EXTENSIONES.md §6.

Incluye a propósito el caso de derivación de intervalo ILAC del historial
real (2 calibraciones ~6 meses de separación, sin intervalo_adoptado_meses
declarado) - es el punto de diseño más importante de esta fase.

Uso:
    python generar_excel_prueba_fase5_limpio.py [ruta_salida.xlsx] [sufijo]
"""
import sys
import openpyxl

RUTA = sys.argv[1] if len(sys.argv) > 1 else "excel_fase5_limpio.xlsx"
SUF = sys.argv[2] if len(sys.argv) > 2 else ""

EQ1, EQ2 = f"EQ-F5-OK-001{SUF}", f"EQ-F5-OK-002{SUF}"
CAL1, CAL2 = f"CAL-F5-OK-001{SUF}", f"CAL-F5-OK-002{SUF}"
VER1, VER2 = f"VER-F5-OK-001{SUF}", f"VER-F5-OK-002{SUF}"

wb = openpyxl.Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("Equipos")
ws.append(["codigo", "nombre", "estado"])
ws.append([EQ1, "Balanza Fase 5 limpia", "operativo"])
ws.append([EQ2, "Termómetro Fase 5 limpio", "operativo"])

ws = wb.create_sheet("Magnitudes")
ws.append(["codigo_equipo", "nombre_magnitud", "emp_valor"])
ws.append([EQ1, "Masa", 1.0])
ws.append([EQ2, "Temperatura", 0.5])

ws = wb.create_sheet("Calibraciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha_calibracion"])
# ~6 meses de separación real -> usado para derivar el intervalo adoptado
# de la Evaluación ILAC de EQ1/Masa (ver PLAN_FASE5_EXTENSIONES.md §1.3)
ws.append([CAL1, EQ1, "Masa", "2022-01-01"])
ws.append([CAL2, EQ1, "Masa", "2022-07-01"])

ws = wb.create_sheet("PuntosCalibracion")
ws.append(["id_temporal_calibracion", "numero_punto", "valor_patron", "valor_indicado"])
ws.append([CAL1, 1, 10, 10.2])
ws.append([CAL2, 1, 10, 10.1])

ws = wb.create_sheet("PlanesVerificacion")
ws.append(["codigo_equipo", "nombre_magnitud", "frecuencia_meses"])
ws.append([EQ1, "Masa", 6])
ws.append([EQ2, "Temperatura", 12])

ws = wb.create_sheet("Verificaciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha"])
ws.append([VER1, EQ1, "Masa", "2023-01-01"])
ws.append([VER2, EQ2, "Temperatura", "2023-02-01"])

ws = wb.create_sheet("PuntosVerificacion")
ws.append(["id_temporal_verificacion", "numero_punto", "valor_patron", "valor_indicado"])
ws.append([VER1, 1, 10, 10.1])
ws.append([VER2, 1, 20, 20.1])

ws = wb.create_sheet("Evaluaciones")
ws.append(["codigo_equipo", "nombre_magnitud"])
# Sin intervalo_adoptado_meses declarado a propósito, en ambas filas:
# EQ1/Masa debe derivar ~6 (del historial real de arriba); EQ2/Temperatura
# no tiene historial suficiente -> debe igualar al sugerido (sin exceso)
ws.append([EQ1, "Masa"])
ws.append([EQ2, "Temperatura"])

ws = wb.create_sheet("PlanesMantenimiento")
ws.append(["codigo_equipo", "tipo", "frecuencia_meses"])
ws.append([EQ1, "preventivo", 6])

ws = wb.create_sheet("Mantenimientos")
ws.append(["codigo_equipo", "tipo", "origen", "titulo"])
ws.append([EQ1, "correctivo", "interno", "Cambio de sensor"])

wb.save(RUTA)
print(f"Archivo Fase 5 LIMPIO generado: {RUTA} (sufijo: '{SUF}')")
print("Debe dar 0 desviaciones en --dry-run.")
