#!/usr/bin/env python3
"""
generar_excel_prueba_migracion.py — genera un archivo .xlsx en la plantilla
estándar de importación (docs/migracion/PLAN_IMPORTACION_EXCEL.md §3), con
una mezcla de filas correctas y errores DELIBERADOS, uno por cada regla de
validación que documenta docs/migracion/PLAN_IMPORTACION_EXCEL.md §4-§5 y
GUIA_VALIDACION_Y_DESVIACIONES.md §3.

Es la Fase 4 del plan: "Prueba con un Excel de prueba que contenga errores
deliberados... confirmar que el modo --dry-run los detecta todos, antes de
escribir tests pytest formales".

Uso:
    python generar_excel_prueba_migracion.py [ruta_salida.xlsx]
"""
import sys
import openpyxl

RUTA = sys.argv[1] if len(sys.argv) > 1 else "excel_prueba_desviaciones.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# -- Hoja Equipos ---------------------------------------------------------
ws = wb.create_sheet("Equipos")
ws.append(["codigo", "nombre", "descripcion", "marca", "modelo", "numero_serie",
           "numero_inventario", "fecha_adquisicion", "costo", "area", "ubicacion",
           "responsable", "estado"])
ws.append(["EQ-001", "Balanza analítica", "", "Mettler Toledo", "XS205", "SN-001",
           "", "2020-01-15", 8500000, "Laboratorio 1", "Mesa 3", "Juan Pérez", "operativo"])
ws.append(["EQ-002", "Termómetro digital", "", "Fluke", "1523", "SN-002",
           "", "2019-06-01", 1200000, "Laboratorio 1", "Estante A", "Ana Gómez", "operativo"])
# Error deliberado 1: código EQ-002 repetido -> Capa 1, Crítica
ws.append(["EQ-002", "Termómetro digital (copia)", "", "Fluke", "1523", "SN-003",
           "", "2019-06-01", 1200000, "Laboratorio 1", "Estante A", "Ana Gómez", "operativo"])
# Error deliberado 2: nombre vacío (obligatorio) -> nivel 1, Crítica
ws.append(["EQ-004", None, "", "", "", "", "", "", None, "", "", "", ""])
# Error deliberado 3: fecha_adquisicion no parseable (opcional) -> nivel 2, Media
ws.append(["EQ-005", "Manómetro", "", "Wika", "233.50", "SN-005",
           "", "no es una fecha", None, "Planta", "", "", "operativo"])
# Error deliberado 4: estado no válido -> nivel 2, Media
ws.append(["EQ-006", "Higrómetro", "", "Testo", "608", "SN-006",
           "", "2021-03-10", None, "Laboratorio 2", "", "", "en_reparacion"])

# -- Hoja Magnitudes -------------------------------------------------------
ws = wb.create_sheet("Magnitudes")
ws.append(["codigo_equipo", "nombre_magnitud", "simbolo", "unidad", "rango_min",
           "rango_max", "resolucion", "emp_texto", "emp_valor", "emp_unidad",
           "clase_exactitud", "tipo_instrumento"])
ws.append(["EQ-001", "Masa", "m", "g", 0, 220, "0.0001", "±0.5 mg", 0.5, "mg", "I", "continuo"])
# Error deliberado 5: sin emp_valor -> nivel 2, Media (semáforo no calculable)
ws.append(["EQ-002", "Temperatura", "T", "°C", -20, 100, "0.1", "", None, "", "", "continuo"])
# Error deliberado 6: codigo_equipo inexistente (EQ-999 no está en Equipos) -> nivel 3, Crítica
ws.append(["EQ-999", "Presión", "P", "kPa", 0, 500, "0.1", "±1 kPa", 1.0, "kPa", "", "continuo"])

# -- Hoja Calibraciones -----------------------------------------------------
ws = wb.create_sheet("Calibraciones")
ws.append(["id_temporal", "codigo_equipo", "nombre_magnitud", "fecha_calibracion",
           "numero_certificado", "laboratorio", "acreditacion_laboratorio",
           "proxima_calibracion", "patrones_utilizados", "metodo_calibracion",
           "temperatura_ambiente", "humedad_relativa", "trazabilidad",
           "observaciones", "costo", "resultado"])
# Caso limpio, pero con puntos que van a estar FUERA de EMP (ver hoja de puntos)
# y resultado declarado "aprobado" -> nivel 4, Alta (semáforo no coincide)
ws.append(["CAL-001", "EQ-001", "Masa", "2023-01-15", "CERT-1001", "Lab Metrología SAS",
           "ONAC", "2024-01-15", "Pesas patrón clase E2", "Comparación directa",
           22.5, 45, "Trazable a INM", "", 250000, "aprobado"])
ws.append(["CAL-002", "EQ-002", "Temperatura", "2023-02-10", "CERT-1002", "Lab Térmico Ltda",
           "ONAC", "2024-02-10", "Baño termostático", "Comparación directa",
           23.0, 50, "Trazable a INM", "", 180000, ""])
# Error deliberado 7: id_temporal repetido (CAL-001 ya existe arriba) -> Capa 1, Crítica
ws.append(["CAL-001", "EQ-002", "Temperatura", "2023-03-01", "CERT-1003", "Lab Térmico Ltda",
           "ONAC", "2024-03-01", "Baño termostático", "Comparación directa",
           23.0, 50, "", "", 180000, ""])
# Error deliberado 8: magnitud inexistente para ese equipo (EQ-001 no tiene "Presión") -> nivel 3, Crítica
ws.append(["CAL-004", "EQ-001", "Presión", "2023-04-01", "CERT-1004", "Lab X",
           "", "", "", "", None, None, "", "", None, ""])
# Error deliberado 9: fecha_calibracion no parseable (obligatoria) -> nivel 2, Crítica
ws.append(["CAL-005", "EQ-002", "Temperatura", "no es fecha", "CERT-1005", "Lab Térmico Ltda",
           "", "", "", "", None, None, "", "", None, ""])
# Error deliberado 10: fecha muy antigua, sospechosa pero válida -> nivel 2, Baja
ws.append(["CAL-006", "EQ-002", "Temperatura", "1985-05-20", "CERT-1006", "Lab Térmico Ltda",
           "", "", "", "", None, None, "", "", None, ""])
# Error deliberado 11: fecha futura, sospechosa pero válida -> nivel 2, Baja
ws.append(["CAL-007", "EQ-002", "Temperatura", "2030-01-01", "CERT-1007", "Lab Térmico Ltda",
           "", "", "", "", None, None, "", "", None, ""])

# -- Hoja PuntosCalibracion -------------------------------------------------
ws = wb.create_sheet("PuntosCalibracion")
ws.append(["id_temporal_calibracion", "numero_punto", "valor_patron", "valor_indicado",
           "incertidumbre", "observacion"])
# Puntos de CAL-001 (EQ-001/Masa, emp_valor=0.5): error = 0.6 en ambos -> fuera de EMP
# a propósito, para que choque con "resultado=aprobado" declarado arriba (nivel 4, Alta)
ws.append(["CAL-001", 1, 10, 10.6, 0.05, ""])
ws.append(["CAL-001", 2, 20, 20.6, 0.05, ""])
# Puntos de CAL-002 (EQ-002/Temperatura, sin EMP -> no se puede evaluar nivel 4, previsto)
ws.append(["CAL-002", 1, 25.0, 25.1, 0.1, ""])
ws.append(["CAL-002", 2, 50.0, 50.2, 0.1, ""])
# Error deliberado 12: referencia a una calibración que no existe -> nivel 3, Crítica
ws.append(["CAL-999", 1, 10, 10.1, 0.05, ""])
# Error deliberado 13: valor_indicado no numérico (obligatorio) -> nivel 2, Crítica
ws.append(["CAL-006", 1, 15, "no es un número", 0.1, ""])

wb.save(RUTA)
print(f"Archivo de prueba generado: {RUTA}")
print("\nErrores deliberados incluidos (13 en total) — ver comentarios en este script "
      "para el detalle de cada uno y qué severidad/regla debería disparar.")
